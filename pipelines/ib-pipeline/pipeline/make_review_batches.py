"""Emit compact review batches from a review xlsx for agent review.

Mechanical only: formats rows into a compact pipe-delimited line per profile.
No filtering or judgment — that's the reviewing agent's job.

Usage:
  python3 make_review_batches.py ../reviews/2026-07_tranche1/search_a_review.xlsx \
      --out-dir ../reviews/2026-07_tranche1/search_a_batches --batch-size 40
"""
import argparse
from pathlib import Path
from openpyxl import load_workbook


def fmt_row(i, r, idx):
    def v(col):
        return r[idx[col]] if r[idx[col]] not in (None, "") else "None"
    return (f"{i}\t{v('linkedin_url')} ||| {v('name')} ||| {v('current_title')} ||| "
            f"{v('current_company')} ||| schools: {v('school_names')} ||| "
            f"headline: {v('headline')} ||| tenure: {v('tenure_current')} ||| "
            f"exp: {v('experience_summary')} ||| src: {v('found_by')}")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("review_sheet")
    ap.add_argument("--out-dir", required=True)
    ap.add_argument("--batch-size", type=int, default=40)
    ap.add_argument("--scope-only", action="store_true",
                    help="only emit rows where review_scope == REVIEW (Search-B scored sheets)")
    args = ap.parse_args()

    wb = load_workbook(args.review_sheet, read_only=True)
    ws = wb["review"]
    header = [c.value for c in ws[1]]
    idx = {h: i for i, h in enumerate(header)}
    rows = [r for r in ws.iter_rows(min_row=2, values_only=True) if r and any(r)]
    if args.scope_only:
        rows = [r for r in rows if r[idx["review_scope"]] == "REVIEW"]

    out = Path(args.out_dir)
    out.mkdir(parents=True, exist_ok=True)
    n_batches = 0
    for b, start in enumerate(range(0, len(rows), args.batch_size)):
        chunk = rows[start:start + args.batch_size]
        lines = [fmt_row(i + 1, r, idx) for i, r in enumerate(chunk)]
        (out / f"batch_{b}.txt").write_text("\n".join(lines))
        n_batches += 1
    print(f"{len(rows)} profiles -> {n_batches} batches of <= {args.batch_size} in {out}")


if __name__ == "__main__":
    main()

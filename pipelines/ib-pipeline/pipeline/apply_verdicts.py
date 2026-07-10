"""Merge agent verdict batches back into a review xlsx by linkedin_url.

Mechanical only: applies verdicts verbatim as written by reviewing agents.
Makes no judgment of its own.

Usage:
  python3 apply_verdicts.py ../reviews/2026-07_tranche1/search_a_review.xlsx \
      "../reviews/2026-07_tranche1/search_a_batches/verdicts_batch_*.txt"
"""
import argparse
import glob
from pathlib import Path
from openpyxl import load_workbook


def parse_verdict_line(line):
    parts = [p.strip() for p in line.split("|||")]
    if len(parts) < 5:
        return None
    url = parts[0].split("\t")[-1].strip()
    return {
        "linkedin_url": url,
        "review_verdict": parts[1].upper(),
        "review_reason": parts[2],
        "verified_school": parts[3],
        "verified_persona": parts[4],
    }


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("review_sheet")
    ap.add_argument("verdict_globs", nargs="+")
    args = ap.parse_args()

    verdicts = {}
    for pattern in args.verdict_globs:
        for f in sorted(glob.glob(pattern)):
            for line in Path(f).read_text().splitlines():
                if not line.strip():
                    continue
                v = parse_verdict_line(line)
                if v:
                    verdicts[v["linkedin_url"]] = v
    print(f"{len(verdicts)} verdicts loaded")

    wb = load_workbook(args.review_sheet)
    ws = wb["review"]
    header = [c.value for c in ws[1]]
    idx = {h: i for i, h in enumerate(header)}
    applied = missing = 0
    for row in ws.iter_rows(min_row=2):
        url = row[idx["linkedin_url"]].value
        if not url:
            continue
        v = verdicts.get(str(url).strip())
        if not v:
            missing += 1
            continue
        row[idx["review_verdict"]].value = v["review_verdict"]
        row[idx["review_reason"]].value = v["review_reason"]
        row[idx["verified_school"]].value = v["verified_school"]
        row[idx["verified_persona"]].value = v["verified_persona"]
        applied += 1
    wb.save(args.review_sheet)
    print(f"applied {applied}, {missing} rows had no matching verdict")


if __name__ == "__main__":
    main()
